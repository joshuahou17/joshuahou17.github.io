#!/usr/bin/env python3
"""
Parse Bible Reading Plan from Excel → JSON.

Reads 'Bible Reading Plan from Joshua.xlsx', sheet '365-Day Plan',
and outputs reading_plan.json with normalized passage references,
OSIS IDs, and genre assignments.
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "Bible Reading Plan from Joshua.xlsx"
OUTPUT_PATH = Path(__file__).parent / "reading_plan.json"

# Column layout: (passage_col, date_col) for each weekday
DAY_COLUMNS = [
    ("E", "F", "Sunday", "Epistles"),
    ("I", "J", "Monday", "The Law"),
    ("M", "N", "Tuesday", "History"),
    ("Q", "R", "Wednesday", "Psalms"),
    ("U", "V", "Thursday", "Poetry"),
    ("Y", "Z", "Friday", "Prophecy"),
    ("AC", "AD", "Saturday", "Gospels"),
]

# Abbreviation → Full book name
BOOK_NAMES = {
    "Gen": "Genesis", "Ex": "Exodus", "Lev": "Leviticus", "Num": "Numbers",
    "Deut": "Deuteronomy", "Josh": "Joshua", "Judg": "Judges", "Ruth": "Ruth",
    "1Sam": "1 Samuel", "2Sam": "2 Samuel", "1Ki": "1 Kings", "2Ki": "2 Kings",
    "1Chr": "1 Chronicles", "2Chr": "2 Chronicles", "Ezra": "Ezra", "Neh": "Nehemiah",
    "Esther": "Esther", "Job": "Job", "Ps": "Psalms", "Prov": "Proverbs",
    "Eccl": "Ecclesiastes", "Song": "Song of Solomon",
    "Isa": "Isaiah", "Jer": "Jeremiah", "Lamentations": "Lamentations",
    "Ezek": "Ezekiel", "Dan": "Daniel",
    "Hosea": "Hosea", "Joel": "Joel", "Amos": "Amos", "Obadiah": "Obadiah",
    "Jonah": "Jonah", "Micah": "Micah", "Nahum": "Nahum", "Habakkuk": "Habakkuk",
    "Zephaniah": "Zephaniah", "Haggai": "Haggai", "Zechariah": "Zechariah",
    "Malachi": "Malachi",
    "Matt": "Matthew", "Mark": "Mark", "Luke": "Luke", "John": "John",
    "Acts": "Acts", "Rom": "Romans", "1Cor": "1 Corinthians", "2Cor": "2 Corinthians",
    "Gal": "Galatians", "Eph": "Ephesians", "Phil": "Philippians", "Col": "Colossians",
    "1Thes": "1 Thessalonians", "2Thes": "2 Thessalonians",
    "1Tim": "1 Timothy", "2Tim": "2 Timothy", "Titus": "Titus", "Philemon": "Philemon",
    "Heb": "Hebrews", "James": "James", "1Pet": "1 Peter", "2Pet": "2 Peter",
    "1John": "1 John", "2John": "2 John", "3John": "3 John", "Jude": "Jude",
    "Rev": "Revelation",
}

# Full book name → OSIS abbreviation (for API.Bible)
OSIS_BOOK_MAP = {
    "Genesis": "GEN", "Exodus": "EXO", "Leviticus": "LEV", "Numbers": "NUM",
    "Deuteronomy": "DEU", "Joshua": "JOS", "Judges": "JDG", "Ruth": "RUT",
    "1 Samuel": "1SA", "2 Samuel": "2SA", "1 Kings": "1KI", "2 Kings": "2KI",
    "1 Chronicles": "1CH", "2 Chronicles": "2CH", "Ezra": "EZR", "Nehemiah": "NEH",
    "Esther": "EST", "Job": "JOB", "Psalms": "PSA", "Proverbs": "PRO",
    "Ecclesiastes": "ECC", "Song of Solomon": "SNG",
    "Isaiah": "ISA", "Jeremiah": "JER", "Lamentations": "LAM",
    "Ezekiel": "EZK", "Daniel": "DAN",
    "Hosea": "HOS", "Joel": "JOL", "Amos": "AMO", "Obadiah": "OBA",
    "Jonah": "JON", "Micah": "MIC", "Nahum": "NAM", "Habakkuk": "HAB",
    "Zephaniah": "ZEP", "Haggai": "HAG", "Zechariah": "ZEC", "Malachi": "MAL",
    "Matthew": "MAT", "Mark": "MRK", "Luke": "LUK", "John": "JHN",
    "Acts": "ACT", "Romans": "ROM", "1 Corinthians": "1CO", "2 Corinthians": "2CO",
    "Galatians": "GAL", "Ephesians": "EPH", "Philippians": "PHP", "Colossians": "COL",
    "1 Thessalonians": "1TH", "2 Thessalonians": "2TH",
    "1 Timothy": "1TI", "2 Timothy": "2TI", "Titus": "TIT", "Philemon": "PHM",
    "Hebrews": "HEB", "James": "JAS", "1 Peter": "1PE", "2 Peter": "2PE",
    "1 John": "1JN", "2 John": "2JN", "3 John": "3JN", "Jude": "JUD",
    "Revelation": "REV",
}

# Single-chapter books (entire book is one chapter)
SINGLE_CHAPTER_BOOKS = {
    "Obadiah", "Philemon", "2 John", "3 John", "Jude",
}

# Chapter counts for multi-chapter books that appear as whole-book references
BOOK_CHAPTER_COUNTS = {
    "Ruth": 4, "Joel": 3, "Jonah": 4, "Nahum": 3, "Habakkuk": 3,
    "Zephaniah": 3, "Haggai": 2, "Malachi": 4, "Micah": 7,
    "Lamentations": 5, "2 Peter": 3, "2 Thessalonians": 3,
    "Titus": 3, "2 John": 1, "3 John": 1, "Jude": 1, "Obadiah": 1,
    "Philemon": 1,
}


def normalize_passage(raw: str) -> dict:
    """Parse a raw passage string into normalized components.

    Examples:
        "Rom 1-2" → {"passage": "Romans 1-2", "book": "Romans", "chapters": [1, 2]}
        "3John"   → {"passage": "3 John", "book": "3 John", "chapters": [1]}
        "Ps 119"  → {"passage": "Psalms 119", "book": "Psalms", "chapters": [119]}
    """
    raw = raw.strip()

    # Match: optional number prefix + book name + optional chapter range
    m = re.match(r'^(\d?\s*[A-Za-z]+)\s*(.*)$', raw)
    if not m:
        raise ValueError(f"Cannot parse passage: {raw}")

    abbrev = m.group(1).strip()
    chapter_part = m.group(2).strip()

    # Look up full book name
    full_name = BOOK_NAMES.get(abbrev)
    if not full_name:
        raise ValueError(f"Unknown book abbreviation: {abbrev} (from '{raw}')")

    # Parse chapters
    if not chapter_part:
        # Whole-book reference — resolve to actual chapter range
        if full_name in BOOK_CHAPTER_COUNTS:
            count = BOOK_CHAPTER_COUNTS[full_name]
            chapters = list(range(1, count + 1))
        else:
            chapters = [1]  # fallback
        passage = full_name
    elif "-" in chapter_part:
        start, end = chapter_part.split("-")
        chapters = list(range(int(start), int(end) + 1))
        passage = f"{full_name} {chapter_part}"
    else:
        chapters = [int(chapter_part)]
        passage = f"{full_name} {chapter_part}"

    return {
        "passage": passage,
        "book": full_name,
        "chapters": chapters,
    }


def build_osis_id(book: str, chapters: list[int]) -> str:
    """Build OSIS passage ID for API.Bible.

    Examples:
        ("Romans", [1, 2]) → "ROM.1-ROM.2"
        ("3 John", [1])    → "3JN.1"
        ("Joel", [])       → "JOL.1-JOL.3"
    """
    osis = OSIS_BOOK_MAP.get(book)
    if not osis:
        raise ValueError(f"No OSIS mapping for: {book}")

    if not chapters:
        # Whole book — use intro reference
        return osis
    elif len(chapters) == 1:
        return f"{osis}.{chapters[0]}"
    else:
        return f"{osis}.{chapters[0]}-{osis}.{chapters[-1]}"


def parse_excel(excel_path: Path) -> list[dict]:
    """Parse the Excel reading plan into a list of day entries."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["365-Day Plan"]

    entries = []
    day_number = 0

    for row_num in range(8, 60):  # Rows 8-59 = weeks 1-52
        week = ws[f"C{row_num}"].value
        if week is None:
            continue

        for passage_col, date_col, weekday, genre in DAY_COLUMNS:
            passage_raw = ws[f"{passage_col}{row_num}"].value
            date_val = ws[f"{date_col}{row_num}"].value

            if not passage_raw or not date_val:
                continue

            day_number += 1
            passage_raw = str(passage_raw).strip()

            # Parse date
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            # Normalize passage
            parsed = normalize_passage(passage_raw)

            # Build OSIS ID
            osis_id = build_osis_id(parsed["book"], parsed["chapters"])

            entries.append({
                "day_number": day_number,
                "date": date_str,
                "weekday": weekday,
                "genre": genre,
                "passage": parsed["passage"],
                "passage_raw": passage_raw,
                "book": parsed["book"],
                "chapters": parsed["chapters"],
                "osis_id": osis_id,
                "week": int(week),
            })

    return entries


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_PATH

    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Parsing: {excel_path}")
    entries = parse_excel(excel_path)
    print(f"Parsed {len(entries)} days")

    # Validate
    dates = [e["date"] for e in entries]
    assert entries[0]["date"] == "2026-03-29", f"Expected start date 2026-03-29, got {entries[0]['date']}"
    assert entries[0]["passage"] == "Romans 1-2", f"Expected Day 1 = Romans 1-2, got {entries[0]['passage']}"
    assert len(set(dates)) == len(dates), "Duplicate dates found"
    print(f"Date range: {dates[0]} to {dates[-1]}")
    print(f"Day 1: {entries[0]['passage']} ({entries[0]['genre']})")
    print(f"Last day: {entries[-1]['passage']} ({entries[-1]['genre']})")

    # Write JSON
    OUTPUT_PATH.write_text(json.dumps(entries, indent=2), encoding="utf-8")
    print(f"Written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
